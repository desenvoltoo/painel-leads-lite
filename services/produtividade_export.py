# -*- coding: utf-8 -*-
"""Exportação XLSX da produtividade da equipe sem dependência de lotes."""
from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict
import unicodedata

from flask import jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import database as db

EXPORT_HEADERS = [
    "status_inscricao", "data_inscricao", "origem", "unidade", "tipo_negocio",
    "curso", "modalidade", "turno", "nome", "cpf", "celular", "email",
    "data_ultima_acao", "qtd_acionamentos", "status", "data_disparo",
    "peca_disparo", "texto_disparo", "consultor_disparo", "tipo_disparo",
    "campanha", "observacao", "data_matricula", "matriculado", "canal",
    "acao_comercial", "consultor_comercial",
]

INVALID_CONSULTANTS = {
    "", "N", "NULL", "N/A", "NA", "NONE", "UNDEFINED", "-",
    "SEM CONSULTOR", "SEM_CONSULTOR", "SEM RESPONSAVEL", "SEM RESPONSÁVEL",
    "NAO INFORMADO", "NÃO INFORMADO",
}


def _rows(sql: str, params: Dict[str, Any] | None = None, name: str = "export_produtividade"):
    return db._run_gestao_query(sql, params or {}, name)


def _relation() -> str:
    schema_raw = (getattr(db, "DB_SCHEMA", None) or "modelo_estrela").strip()
    schema_ident = db._safe_ident(schema_raw)
    found = _rows(
        """
        SELECT table_name
        FROM (
            SELECT table_name, 0 AS prioridade
            FROM information_schema.views
            WHERE table_schema = :schema
              AND table_name = 'vw_leads_painel_lite'
            UNION ALL
            SELECT table_name, 1 AS prioridade
            FROM information_schema.views
            WHERE table_schema = :schema
              AND table_name = 'leads_painel_lite'
            UNION ALL
            SELECT table_name, 2 AS prioridade
            FROM information_schema.tables
            WHERE table_schema = :schema
              AND table_name = 'leads_painel_lite'
        ) origem
        ORDER BY prioridade
        LIMIT 1
        """,
        {"schema": schema_raw},
        "export_produtividade_relation",
    )
    if not found:
        raise RuntimeError("View de leads não encontrada no banco.")
    return f"{schema_ident}.{db._safe_ident(str(found[0]['table_name']))}"


def _parse_month(value: str | None) -> tuple[date, date, str]:
    raw = (value or datetime.now().strftime("%Y-%m")).strip()
    try:
        year, month = [int(part) for part in raw.split("-", 1)]
        start = date(year, month, 1)
        end = date(year + (month == 12), 1 if month == 12 else month + 1, 1)
    except Exception as exc:
        raise ValueError("Mês inválido. Use AAAA-MM.") from exc
    return start, end, f"{month:02d}/{year}"


def _normalize(value: Any) -> str:
    text = str(value or "").strip().upper()
    return "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )


def _valid_consultant(value: Any) -> bool:
    normalized = _normalize(str(value or "").replace("\\", ""))
    return normalized not in {_normalize(item) for item in INVALID_CONSULTANTS}


def _is_positive(row: Dict[str, Any]) -> bool:
    text = " ".join(_normalize(row.get(key)) for key in ("status", "observacao", "acao_comercial"))
    return any(token in text for token in ("POSIT", "INTERESS", "MATRIC", "CONVERT", "FECHOU"))


def _is_matriculated(row: Dict[str, Any]) -> bool:
    # Regra oficial: somente a flag booleana verdadeira confirma matrícula.
    return row.get("matriculado") is True


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value in (None, ""):
        return None

    text = str(value).strip()
    if not text:
        return None

    # PostgreSQL/SQLAlchemy normalmente retorna ISO 8601 com T, microssegundos
    # e, às vezes, timezone. datetime.fromisoformat cobre todos esses casos.
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        pass

    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass

    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


def _category_date(row: Dict[str, Any], category: str) -> date | None:
    if category == "MATRÍCULAS":
        return _as_date(row.get("data_matricula"))
    return _as_date(row.get("data_disparo"))


def _matches_category(row: Dict[str, Any], category: str) -> bool:
    kind = _normalize(row.get("tipo_disparo"))
    if category == "URA":
        return "URA" in kind
    if category == "ROBO":
        return "ROBO" in kind
    if category == "RETORNO POSITIVO / URA":
        return "URA" in kind and _is_positive(row)
    if category == "RETORNO POSITIVO / ROBO":
        return "ROBO" in kind and _is_positive(row)
    if category == "MATRÍCULAS":
        return _is_matriculated(row)
    return False


def _in_export_month(value: Any, start: date) -> bool:
    parsed = _as_date(value)
    return bool(parsed and parsed.year == start.year and parsed.month == start.month)


def _make_workbook(rows: list[Dict[str, Any]], start: date, month_label: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "FECHAMENTO"
    leads_ws = wb.create_sheet("FECHARAM MATRICULA")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_blue = PatternFill("solid", fgColor="B4C6E7")
    fill_yellow = PatternFill("solid", fgColor="FFD966")
    fill_green = PatternFill("solid", fgColor="00B050")
    fill_red = PatternFill("solid", fgColor="FF0000")
    fill_orange = PatternFill("solid", fgColor="C65911")
    white_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    days = monthrange(start.year, start.month)[1]
    categories = ["URA", "ROBO", "RETORNO POSITIVO / URA", "RETORNO POSITIVO / ROBO", "MATRÍCULAS"]
    by_consultant: dict[str, list[Dict[str, Any]]] = {}
    for row in rows:
        consultant = str(row.get("consultor_disparo") or "").strip().upper()
        if not _valid_consultant(consultant):
            continue
        by_consultant.setdefault(consultant, []).append(row)

    current_row = 1
    for consultant in sorted(by_consultant):
        data = by_consultant[consultant]
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=days + 3)
        title = ws.cell(current_row, 4, "ANÁLISE DESEMPENHO ROBO")
        title.fill, title.font, title.alignment = fill_blue, bold, Alignment(horizontal="center")
        current_row += 1

        ws.cell(current_row, 3, "Disparos realizados").fill = fill_yellow
        ws.cell(current_row, 3).font = bold
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=days + 3)
        month_cell = ws.cell(current_row, 4, month_label.upper())
        month_cell.fill, month_cell.font, month_cell.alignment = fill_yellow, bold, Alignment(horizontal="center")
        current_row += 2

        ws.cell(current_row, 1, consultant).fill = fill_yellow
        ws.cell(current_row, 1).font = bold
        ws.cell(current_row, 2, "TOTAL").fill = fill_yellow
        ws.cell(current_row, 2).font = bold
        for day in range(1, days + 1):
            cell = ws.cell(current_row, day + 3, day)
            weekday = date(start.year, start.month, day).weekday()
            cell.fill = fill_red if weekday >= 5 else fill_green
            cell.font = white_font if weekday >= 5 else bold
            cell.alignment, cell.border = Alignment(horizontal="center"), border
        header_row = current_row
        current_row += 1

        for category in categories:
            ws.cell(current_row, 1, category).font = bold
            daily = []
            for day in range(1, days + 1):
                count = 0
                for row in data:
                    event_date = _category_date(row, category)
                    if not event_date:
                        continue
                    if event_date.year != start.year or event_date.month != start.month or event_date.day != day:
                        continue
                    if _matches_category(row, category):
                        count += 1
                daily.append(count)
                cell = ws.cell(current_row, day + 3, count if count else None)
                cell.alignment, cell.border = Alignment(horizontal="center"), border
            ws.cell(current_row, 2, sum(daily)).font = bold
            if category == "MATRÍCULAS":
                for col in range(1, days + 4):
                    ws.cell(current_row, col).fill = fill_yellow
            current_row += 1

        total_dispatches = sum(1 for row in data if _in_export_month(row.get("data_disparo"), start))
        ws.cell(header_row, 2, total_dispatches).font = bold
        current_row += 3

    ws.cell(current_row, 1, "Mês").fill = fill_blue
    ws.cell(current_row, 2, month_label.upper()).fill = fill_blue
    current_row += 1
    for col, value in ((1, "EQUIPE"), (2, "TOTAL")):
        ws.cell(current_row, col, value).fill = fill_orange
        ws.cell(current_row, col).font = white_font
    current_row += 1
    total_team = 0
    for consultant in sorted(by_consultant):
        total = sum(
            1 for row in by_consultant[consultant]
            if _is_matriculated(row) and _in_export_month(row.get("data_matricula"), start)
        )
        total_team += total
        ws.cell(current_row, 1, consultant)
        ws.cell(current_row, 2, total)
        current_row += 1
    for col, value in ((1, "TOTAL"), (2, total_team)):
        ws.cell(current_row, col, value).fill = fill_orange
        ws.cell(current_row, col).font = white_font

    ws.freeze_panes = "D1"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 3
    for col in range(4, days + 4):
        ws.column_dimensions[get_column_letter(col)].width = 5

    for col, header in enumerate(EXPORT_HEADERS, 1):
        cell = leads_ws.cell(1, col, header)
        cell.fill, cell.font, cell.border = fill_blue, bold, border
        cell.alignment = Alignment(horizontal="center")

    matriculated = [
        row for row in rows
        if _is_matriculated(row) and _in_export_month(row.get("data_matricula"), start)
    ]
    for r_idx, row in enumerate(matriculated, 2):
        for c_idx, header in enumerate(EXPORT_HEADERS, 1):
            value = row.get(header)
            cell = leads_ws.cell(r_idx, c_idx, value)
            cell.border = border
            if isinstance(value, (datetime, date)):
                cell.number_format = "dd/mm/yyyy"

    leads_ws.freeze_panes = "A2"
    leads_ws.auto_filter.ref = leads_ws.dimensions
    for idx, header in enumerate(EXPORT_HEADERS, 1):
        width = 30 if header in {"nome", "curso", "campanha", "texto_disparo", "observacao"} else min(45, max(12, len(header) + 2))
        leads_ws.column_dimensions[get_column_letter(idx)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def register_produtividade_export(app) -> None:
    if "exportar_produtividade_equipe" in app.view_functions:
        return

    @app.get("/api/gestao/operacional/consultores/exportar")
    def exportar_produtividade_equipe():
        try:
            start, end, month_label = _parse_month(request.args.get("mes"))
            relation = _relation()
            selected = ", ".join(db._safe_ident(col) for col in EXPORT_HEADERS)
            rows = _rows(
                f"""
                SELECT {selected}
                FROM {relation}
                WHERE (
                    data_disparo >= :inicio
                    AND data_disparo < :fim
                ) OR (
                    matriculado IS TRUE
                    AND data_matricula >= :inicio
                    AND data_matricula < :fim
                )
                ORDER BY consultor_disparo, COALESCE(data_disparo, data_matricula), nome
                """,
                {"inicio": start.isoformat(), "fim": end.isoformat()},
                "export_produtividade_rows",
            )
            stream = _make_workbook(list(rows or []), start, month_label)
            filename = f"produtividade_equipe_{start.strftime('%Y_%m')}.xlsx"
            return send_file(
                stream,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        except Exception as exc:
            app.logger.exception("Falha ao exportar produtividade da equipe")
            return jsonify({"ok": False, "error": str(exc)}), 500
