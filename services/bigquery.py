# services/bigquery.py
from __future__ import annotations

import os
import logging
import threading
import time
import csv
import json
import re
import unicodedata
from concurrent.futures import TimeoutError as FuturesTimeoutError
from decimal import Decimal, InvalidOperation
from datetime import timedelta, datetime, timezone
from io import StringIO
from pathlib import Path
from time import perf_counter
from typing import Any, Dict, Iterator, List, Optional, Tuple

import google.auth
from google.cloud import bigquery
from cachetools import TTLCache
from google.api_core.exceptions import (
    GoogleAPICallError,
    BadRequest,
    Forbidden,
    NotFound,
    ServiceUnavailable,
    TooManyRequests,
)

# XLSX
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ============================================================
# CONFIG (ENV + travas)
# ============================================================
GCP_PROJECT_ID = (os.getenv("BIGQUERY_PROJECT_ID") or os.getenv("GCP_PROJECT_ID") or "painel-universidade").strip()
BQ_DATASET = (os.getenv("BIGQUERY_DATASET") or os.getenv("BQ_DATASET") or "modelo_estrela").strip()
BQ_LOCATION = os.getenv("BQ_LOCATION", "").strip()

BQ_STAGING_TABLE = os.getenv("BQ_STAGING_TABLE", "stg_leads_site").strip() or "stg_leads_site"
BQ_PROCEDURE = os.getenv("BQ_PROCEDURE", "sp_import_star_from_site")

# Painel lê somente essa view
BQ_VIEW_LEADS = os.getenv("BQ_VIEW_LEADS", "vw_leads_painel_lite")

# Gestão Operacional usa as views dedicadas e recalcula prioridade máxima por status vazio
BQ_VIEW_GESTAO_LEADS_PRIORIZADOS = os.getenv("BQ_VIEW_GESTAO_LEADS_PRIORIZADOS", "vw_leads_priorizados")
BQ_VIEW_GESTAO_PRODUTIVIDADE = os.getenv("BQ_VIEW_GESTAO_PRODUTIVIDADE", "vw_produtividade_consultores")
BQ_VIEW_GESTAO_OPERACAO = os.getenv("BQ_VIEW_GESTAO_OPERACAO", "vw_operacao_rpa")
GESTAO_CACHE_TTL_SECONDS = int(os.getenv("GESTAO_CACHE_TTL_SECONDS", "120"))
GESTAO_FILA_LIMIT = int(os.getenv("GESTAO_FILA_LIMIT", "500"))

DEFAULT_LIMIT = int(os.getenv("BQ_DEFAULT_LIMIT", "200"))
MAX_LIMIT = int(os.getenv("BQ_MAX_LIMIT", "2000"))
EXPORT_MAX_ROWS = int(os.getenv("BQ_EXPORT_MAX_ROWS", "50000"))
QUERY_TIMEOUT_SECONDS = int(os.getenv("BQ_QUERY_TIMEOUT_SECONDS", "180"))

# Retry
RETRY_MAX_ATTEMPTS = int(os.getenv("BQ_RETRY_MAX_ATTEMPTS", "3"))
RETRY_BASE_DELAY_S = float(os.getenv("BQ_RETRY_BASE_DELAY_S", "2.0"))
RETRY_MAX_DELAY_S = float(os.getenv("BQ_RETRY_MAX_DELAY_S", "30.0"))

# colunas que não devem sofrer comportamento numérico
PHONEISH_COLUMNS = {"cpf", "celular"}

EMPTY_FILTER_TOKEN = "__EMPTY__"


def _current_upload_timestamp() -> datetime:
    return datetime.now(timezone.utc)


def _split_empty_filter(values: List[str]) -> Tuple[List[str], bool]:
    cleaned: List[str] = []
    include_empty = False
    for value in values:
        item = str(value or "").strip()
        if not item:
            continue
        if item == EMPTY_FILTER_TOKEN:
            include_empty = True
            continue
        cleaned.append(item)
    return cleaned, include_empty


def _empty_value_condition(column_expr: str) -> str:
    return f"({column_expr} IS NULL OR TRIM(CAST({column_expr} AS STRING)) = '')"


def _apply_multi_value_filter(
    sql: str,
    params: List[Any],
    *,
    col: str,
    param_name: str,
    values: List[str],
) -> str:
    if not values or not _has_view_col(col):
        return sql

    filled_values, include_empty = _split_empty_filter(values)
    conditions: List[str] = []
    column_expr = f"v.{col}"

    if filled_values:
        conditions.append(f"{column_expr} IN UNNEST(@{param_name})")
        params.append(bigquery.ArrayQueryParameter(param_name, "STRING", filled_values))

    if include_empty:
        conditions.append(_empty_value_condition(column_expr))

    if conditions:
        sql += " AND (" + " OR ".join(conditions) + ")"

    return sql

UPLOAD_COLUMN_ALIASES = {
    "unidade": "unidade",
    "polo": "unidade",
    "campus": "unidade",
    "statusinscricao": "status_inscricao",
    "status inscricao": "status_inscricao",
    "status inscrição": "status_inscricao",
    "datainscricao": "data_inscricao",
    "data inscricao": "data_inscricao",
    "data inscrição": "data_inscricao",
    "tipo negocio": "tipo_negocio",
    "tipo negócio": "tipo_negocio",
    "qtd acionamentos": "qtd_acionamentos",
    "data ultima acao": "data_ultima_acao",
    "data última ação": "data_ultima_acao",
    "data disparo": "data_disparo",
    "peca disparo": "peca_disparo",
    "peça disparo": "peca_disparo",
    "texto disparo": "texto_disparo",
    "consultor disparo": "consultor_disparo",
    "tipo disparo": "tipo_disparo",
    "data matricula": "data_matricula",
    "data matrícula": "data_matricula",
    "acao comercial": "acao_comercial",
    "ação comercial": "acao_comercial",
    "consultor comercial": "consultor_comercial",
}

# ============================================================
# CLIENTES THREAD-SAFE
# Cada thread mantém sua própria instância, evitando condições
# de corrida em ambientes multi-thread (FastAPI / Gunicorn).
# ============================================================
_thread_local = threading.local()
_gestao_cache = TTLCache(maxsize=16, ttl=GESTAO_CACHE_TTL_SECONDS)
_gestao_cache_lock = threading.Lock()


def _bq_location() -> Optional[str]:
    """Retorna a região do dataset para consultar jobs na localização correta."""
    if BQ_LOCATION:
        return BQ_LOCATION

    cache_key = f"dataset_location::{GCP_PROJECT_ID}.{BQ_DATASET}"
    cache = getattr(_thread_local, "schema_cache", None)
    if cache is None:
        cache = {}
        _thread_local.schema_cache = cache
    if cache_key in cache:
        return cache[cache_key]

    try:
        dataset = get_bq_client().get_dataset(f"{GCP_PROJECT_ID}.{BQ_DATASET}")
        location = dataset.location or None
        cache[cache_key] = location
        return location
    except Exception as exc:
        logger.warning("Não foi possível detectar a localização do dataset BigQuery: %s", exc)
        cache[cache_key] = None
        return None


def get_bq_client() -> bigquery.Client:
    if not getattr(_thread_local, "bq_client", None):
        _thread_local.bq_client = bigquery.Client(project=GCP_PROJECT_ID)
    return _thread_local.bq_client



def _export_jobs_table_id() -> str:
    return f"{GCP_PROJECT_ID}.{BQ_DATASET}.export_jobs"


def _ensure_export_jobs_table() -> None:
    client = get_bq_client()
    table_id = _export_jobs_table_id()
    try:
        _with_retry(client.get_table, table_id, operation_name="get export_jobs table")
        return
    except NotFound:
        pass
    except Exception:
        logger.exception("Falha ao validar tabela export_jobs")
        raise

    schema = [
        bigquery.SchemaField("job_id", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("status", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("created_at", "TIMESTAMP", mode="NULLABLE"),
        bigquery.SchemaField("updated_at", "TIMESTAMP", mode="NULLABLE"),
        bigquery.SchemaField("metadata", "STRING", mode="NULLABLE"),
    ]
    table = bigquery.Table(table_id, schema=schema)
    _with_retry(client.create_table, table, operation_name="create export_jobs table")


def create_export_job(job_id: str, metadata_dict: Dict[str, Any]) -> None:
    _ensure_export_jobs_table()
    client = get_bq_client()
    now = datetime.utcnow()
    payload = metadata_dict or {}
    row = {
        "job_id": job_id,
        "status": str(payload.get("status") or "queued"),
        "created_at": now,
        "updated_at": now,
        "metadata": json.dumps(payload, ensure_ascii=False, default=str),
    }
    errors = _with_retry(
        client.insert_rows_json,
        _export_jobs_table_id(),
        [row],
        operation_name="insert export job",
    )
    if errors:
        raise RuntimeError(f"Falha ao inserir export job no BigQuery: {errors}")


def _run_query(query: str, params: Optional[List[Any]] = None) -> Any:
    return _run_gestao_query(query, params=params, operation_name="export_job")


def update_export_job(job_id: str, **kwargs) -> None:
    _ensure_export_jobs_table()
    existing = get_export_job(job_id)
    if not existing:
        return
    metadata = dict(existing)
    metadata.update(kwargs)
    status = str(metadata.get("status") or "queued")

    query = f"""
        UPDATE `{_export_jobs_table_id()}`
        SET
          status = @status,
          metadata = @metadata,
          updated_at = CURRENT_TIMESTAMP()
        WHERE job_id = @job_id
    """
    params = [
        bigquery.ScalarQueryParameter("status", "STRING", status),
        bigquery.ScalarQueryParameter("metadata", "STRING", json.dumps(metadata, ensure_ascii=False, default=str)),
        bigquery.ScalarQueryParameter("job_id", "STRING", job_id),
    ]
    client = get_bq_client()
    list(
        client.query(
            query,
            job_config=bigquery.QueryJobConfig(query_parameters=params),
            location=_bq_location(),
        ).result()
    )


def get_export_job(job_id: str) -> Optional[Dict[str, Any]]:
    _ensure_export_jobs_table()
    query = f"""
        SELECT job_id, status, created_at, updated_at, metadata
        FROM `{_export_jobs_table_id()}`
        WHERE job_id = @job_id
        LIMIT 1
    """
    params = [bigquery.ScalarQueryParameter("job_id", "STRING", job_id)]
    client = get_bq_client()
    rows = list(
        client.query(
            query,
            job_config=bigquery.QueryJobConfig(query_parameters=params),
            location=_bq_location(),
        ).result()
    )
    if not rows:
        return None
    row = rows[0]
    metadata_raw = row.get("metadata")
    metadata = {}
    if metadata_raw:
        try:
            metadata = json.loads(metadata_raw)
        except Exception:
            logger.warning("metadata inválido para export job job_id=%s", job_id)
            metadata = {}
    if not isinstance(metadata, dict):
        metadata = {}
    metadata.setdefault("job_id", row.get("job_id"))
    metadata.setdefault("status", row.get("status"))
    metadata.setdefault("created_at", row.get("created_at").isoformat() + "Z" if row.get("created_at") else None)
    metadata.setdefault("updated_at", row.get("updated_at").isoformat() + "Z" if row.get("updated_at") else None)
    return metadata


# ============================================================
# RETRY COM BACKOFF EXPONENCIAL
# Erros fatais (BadRequest, Forbidden, NotFound) não são retriados.
# ============================================================
_RETRYABLE_EXCEPTIONS = (ServiceUnavailable, TooManyRequests, GoogleAPICallError)


def _with_retry(fn, *args, operation_name: str = "operação", **kwargs) -> Any:
    delay = RETRY_BASE_DELAY_S
    last_exc: Optional[Exception] = None
    for attempt in range(1, RETRY_MAX_ATTEMPTS + 1):
        try:
            return fn(*args, **kwargs)
        except (BadRequest, Forbidden, NotFound) as exc:
            logger.error("Erro fatal em '%s' (tentativa %d/%d): %s", operation_name, attempt, RETRY_MAX_ATTEMPTS, exc)
            raise RuntimeError(f"Erro fatal em {operation_name}: {exc}") from exc
        except _RETRYABLE_EXCEPTIONS as exc:
            last_exc = exc
            if attempt == RETRY_MAX_ATTEMPTS:
                break
            logger.warning("Erro transitório em '%s' (tentativa %d/%d), aguardando %.1fs: %s", operation_name, attempt, RETRY_MAX_ATTEMPTS, delay, exc)
            time.sleep(delay)
            delay = min(delay * 2, RETRY_MAX_DELAY_S)
        except Exception as exc:
            logger.exception("Erro inesperado em '%s': %s", operation_name, exc)
            raise
    raise RuntimeError(f"'{operation_name}' falhou após {RETRY_MAX_ATTEMPTS} tentativas.") from last_exc


def _validate_bq_identifier(value: str, label: str) -> str:
    text = str(value or "").strip()
    if not text or not text.replace("_", "a").replace("-", "a").isalnum():
        raise ValueError(f"Identificador BigQuery inválido para {label}.")
    return text


def bq_table_id(name: str) -> str:
    safe_name = _validate_bq_identifier(name, "table_or_view")
    return f"{GCP_PROJECT_ID}.{BQ_DATASET}.{safe_name}"


def bq_ref(name: str) -> str:
    return f"`{bq_table_id(name)}`"


def _tbl(name: str) -> str:
    return bq_ref(name)


def _staging_table_id() -> str:
    return f"{GCP_PROJECT_ID}.{BQ_DATASET}.{BQ_STAGING_TABLE}"


def _view_table_id() -> str:
    return f"{GCP_PROJECT_ID}.{BQ_DATASET}.{BQ_VIEW_LEADS}"


def _view_columns() -> set[str]:
    """Retorna as colunas reais da view/tabela, com cache por thread."""
    cache_key = f"view_columns::{_view_table_id()}"
    cache = getattr(_thread_local, "schema_cache", None)
    if cache is None:
        cache = {}
        _thread_local.schema_cache = cache
    if cache_key not in cache:
        table = get_bq_client().get_table(_view_table_id())
        cache[cache_key] = {field.name for field in table.schema}
    return cache[cache_key]


def _has_view_col(col: str) -> bool:
    return col in _view_columns()


def _table_id(view_name: str) -> str:
    return bq_table_id(view_name)


def _table_columns(view_name: str) -> set[str]:
    """Retorna as colunas reais de uma view/tabela arbitrária, com cache por thread."""
    table_id = _table_id(view_name)
    cache_key = f"table_columns::{table_id}"
    cache = getattr(_thread_local, "schema_cache", None)
    if cache is None:
        cache = {}
        _thread_local.schema_cache = cache
    if cache_key not in cache:
        table = get_bq_client().get_table(table_id)
        cache[cache_key] = {field.name for field in table.schema}
    return cache[cache_key]


def _has_table_col(view_name: str, col: str) -> bool:
    return col in _table_columns(view_name)


def _select_table_col(view_name: str, alias_prefix: str, col: str, alias: Optional[str] = None, bq_type: str = "STRING") -> str:
    """Seleciona coluna de uma view arbitrária se existir; caso contrário devolve NULL tipado."""
    out_alias = alias or col
    if _has_table_col(view_name, col):
        return f"{alias_prefix}.{col} AS {out_alias}"
    return f"CAST(NULL AS {bq_type}) AS {out_alias}"


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        try:
            return float(value)
        except (InvalidOperation, ValueError):
            return str(value)
    if isinstance(value, (datetime, timedelta)):
        return str(value)
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return value


def _rows_to_json_safe(rows: Any) -> List[Dict[str, Any]]:
    return [{key: _json_safe_value(value) for key, value in dict(row).items()} for row in rows]


def _single_row_to_json_safe(rows: Any) -> Dict[str, Any]:
    data = _rows_to_json_safe(rows)
    return data[0] if data else {}


def _select_col(col: str, alias: Optional[str] = None, bq_type: str = "STRING") -> str:
    """Seleciona coluna se existir; caso contrário devolve NULL tipado com o mesmo alias."""
    out_alias = alias or col
    if _has_view_col(col):
        return f"v.{col} AS {out_alias}"
    return f"CAST(NULL AS {bq_type}) AS {out_alias}"


def _first_existing_col(*cols: str) -> Optional[str]:
    existing = _view_columns()
    return next((col for col in cols if col in existing), None)


def _order_expr_for(order_by: str) -> Optional[str]:
    if order_by == "data_inscricao_dt":
        order_by = "data_inscricao"
    allowed_order = {
        "data_inscricao": "data_inscricao",
        "data_disparo": "data_disparo",
        "status": "status",
        "curso": "curso",
        "modalidade": "modalidade",
        "polo": "polo",
        "nome": "nome",
        "cpf": "cpf",
        "canal": "canal",
        "campanha": "campanha",
        "consultor_disparo": "consultor_disparo",
    }
    col = allowed_order.get(order_by, "data_inscricao")
    if _has_view_col(col):
        return f"v.{col}"
    fallback = _first_existing_col("data_inscricao", "data_disparo", "nome", "cpf")
    return f"v.{fallback}" if fallback else None


def _as_list(v: Any) -> List[str]:
    if v is None:
        return []
    if isinstance(v, (list, tuple)):
        return [str(x).strip() for x in v if str(x).strip()]
    s = str(v).strip()
    if not s:
        return []
    if "||" in s:
        return [p.strip() for p in s.split("||") if p.strip()]
    if "," in s:
        return [p.strip() for p in s.split(",") if p.strip()]
    return [s]


def _data_inscricao_order_clause(order_dir: str) -> str:
    """Ordena por data_inscricao, tolerando views sem data_atualizacao."""
    expr = _order_expr_for("data_inscricao")
    if not expr:
        return "1"

    parts = [
        f"CASE WHEN {expr} IS NULL THEN 1 ELSE 0 END ASC",
        f"{expr} {order_dir}",
    ]
    if _has_view_col("data_atualizacao"):
        parts.append("v.data_atualizacao DESC")
    return ",\n    ".join(parts)


def _data_disparo_order_clause(order_dir: str) -> str:
    """Ordena por data_disparo colocando valores vazios primeiro."""
    expr = _order_expr_for("data_disparo")
    if not expr:
        return _data_inscricao_order_clause(order_dir)

    parts = [
        f"CASE WHEN {expr} IS NULL OR TRIM(CAST({expr} AS STRING)) = '' THEN 0 ELSE 1 END ASC",
        f"{expr} {order_dir}",
    ]
    data_inscricao_expr = _order_expr_for("data_inscricao")
    if data_inscricao_expr:
        parts.append(f"{data_inscricao_expr} ASC")
    return ",\n    ".join(parts)


# ============================================================
# NORMALIZADORES
# ============================================================
def _normalize_decimal_string_to_int_string(s: str) -> str:
    """
    Converte representações numéricas integrais para string inteira.

    Exemplos:
      "11974817404.0" -> "11974817404"
      "5.511944391404e13" -> "55119443914040"
      "  12345  " -> "12345"

    Se não for número integral, devolve o valor original.
    """
    s = str(s).strip()
    if not s:
        return s

    try:
        d = Decimal(s)
        if d == d.to_integral_value():
            return str(d.quantize(Decimal("1")))
        return s
    except (InvalidOperation, ValueError):
        return s


def _normalize_phoneish_value(x: Any) -> Optional[str]:
    """
    Normaliza CPF/celular preservando como texto e removendo '.0' indevido.

    Regras:
    - None / NaN / vazio -> None
    - float integral -> string inteira
    - string numérica integral com .0 / notação científica -> string inteira
    - demais strings -> trim simples
    """
    try:
        import pandas as pd
    except Exception:
        pd = None

    if x is None:
        return None

    if pd is not None:
        try:
            if pd.isna(x):
                return None
        except Exception:
            pass

    # números Python
    if isinstance(x, int):
        return str(x)

    if isinstance(x, float):
        if pd is not None:
            try:
                if pd.isna(x):
                    return None
            except Exception:
                pass
        # se for 11974817404.0 -> 11974817404
        if x.is_integer():
            return str(int(x))
        s = str(x).strip()
        return s or None

    s = str(x).strip()
    if not s or s.lower() == "nan":
        return None

    # corrige:
    # "11974817404.0" -> "11974817404"
    # "5.511944391404e13" -> "55119443914040"
    s2 = _normalize_decimal_string_to_int_string(s)

    return s2 or None


def _normalize_column_name(name: Any) -> str:
    """Normaliza cabeçalhos de upload para o schema da staging."""
    raw = str(name or "").strip()
    compact = " ".join(raw.replace("_", " ").replace("-", " ").split()).lower()
    no_accents = "".join(
        ch for ch in unicodedata.normalize("NFKD", compact)
        if not unicodedata.combining(ch)
    )
    snake = re.sub(r"[^a-z0-9]+", "_", no_accents).strip("_")
    return UPLOAD_COLUMN_ALIASES.get(compact) or UPLOAD_COLUMN_ALIASES.get(no_accents) or snake


def _normalize_generic_string(x: Any) -> Optional[str]:
    """
    Converte qualquer valor para STRING segura sem estourar upload.
    """
    try:
        import pandas as pd
    except Exception:
        pd = None

    if x is None:
        return None

    if pd is not None:
        try:
            if pd.isna(x):
                return None
        except Exception:
            pass

    s = str(x).strip()
    if not s or s.lower() == "nan":
        return None
    return s


def _xlsx_safe_cell_value(key: str, value: Any):
    """
    Mantém datas compatíveis com Excel e força CPF/celular como texto.
    """
    from datetime import datetime, date

    if key in PHONEISH_COLUMNS:
        return _normalize_phoneish_value(value)

    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo is not None else value
    if isinstance(value, date):
        return value

    return value


# ============================================================
# STAGING SCHEMA (BLINDADO)
# ============================================================
# Tudo STRING para não quebrar em CSV/XLSX e deixar a SP parsear
STAGING_SCHEMA: List[bigquery.SchemaField] = [
    bigquery.SchemaField("status_inscricao", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("data_inscricao", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("origem", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("unidade", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("tipo_negocio", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("curso", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("modalidade", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("turno", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("nome", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("cpf", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("celular", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("email", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("data_ultima_acao", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("qtd_acionamentos", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("status", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("data_disparo", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("peca_disparo", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("texto_disparo", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("consultor_disparo", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("tipo_disparo", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("campanha", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("observacao", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("data_matricula", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("matriculado", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("canal", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("acao_comercial", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("consultor_comercial", "STRING", mode="NULLABLE"),
    bigquery.SchemaField("dt_upload", "TIMESTAMP", mode="NULLABLE"),
]

STAGING_COLUMNS = tuple(field.name for field in STAGING_SCHEMA)
# A staging aceita todas as colunas como NULLABLE/STRING. Para preservar a
# compatibilidade com uploads parciais existentes, a validação obrigatória fica
# restrita a linhas com algum conteúdo; colunas ausentes são enviadas como NULL.
REQUIRED_UPLOAD_COLUMNS: Tuple[str, ...] = ()
REJECTION_LOG_TABLE = "logs_rejeicoes_import"



def _json_dump_record(record: Dict[str, Any]) -> str:
    return json.dumps(record, default=str, ensure_ascii=False)


def _empty_row_mask(df) -> Any:
    """Retorna máscara booleana para linhas vazias, ignorando metadados do backend."""
    content_columns = [col for col in df.columns if col != "dt_upload"]
    if not content_columns:
        return df.apply(lambda _row: True, axis=1)
    return df[content_columns].apply(
        lambda row: all(value is None or str(value).strip() == "" for value in row),
        axis=1,
    )


def _rejection_log_table_id() -> str:
    return f"{GCP_PROJECT_ID}.{BQ_DATASET}.{REJECTION_LOG_TABLE}"


def _insert_rejection_logs(rejections: List[Dict[str, Any]]) -> None:
    """Registra rejeições por linha sem interromper o upload caso o log falhe."""
    if not rejections:
        return

    client = get_bq_client()
    table_id = _rejection_log_table_id()
    rows = [
        {
            "ts": item["data_hora"],
            "motivo": item["motivo"],
            "cpf_raw": item.get("cpf_raw"),
            "celular_raw": item.get("celular_raw"),
            "nome_raw": item.get("nome_raw"),
            "email_raw": item.get("email_raw"),
            "id_importacao": item.get("id_importacao"),
        }
        for item in rejections
    ]

    try:
        errors = client.insert_rows_json(table_id, rows)
        if errors:
            logger.error("Falha ao inserir logs de rejeição em %s: %s", table_id, errors)
        else:
            logger.info("Logs de rejeição inseridos: tabela=%s linhas=%d", table_id, len(rows))
    except Exception as exc:
        logger.exception("Falha ao registrar rejeições em %s: %s", table_id, exc)


def _build_import_report(
    *,
    filename: str,
    started: float,
    rows_received: int,
    rows_imported: int,
    rows_rejected: int,
    received_columns: List[str],
    rejection_reasons: Dict[str, int],
    staging_rows: Optional[int] = None,
    staging_schema: Optional[List[bigquery.SchemaField]] = None,
    upload_ts: Optional[datetime] = None,
) -> Dict[str, Any]:
    normalized_columns = [_normalize_column_name(c) for c in received_columns]
    staging_columns = tuple(field.name for field in (staging_schema or STAGING_SCHEMA))
    return {
        "arquivo": filename,
        "tabela_destino": _staging_table_id(),
        "linhas_recebidas": rows_received,
        "linhas_importadas": rows_imported,
        "linhas_processadas": rows_imported,
        "linhas_rejeitadas": rows_rejected,
        "linhas_gravadas_staging": staging_rows if staging_rows is not None else rows_imported,
        "motivos_rejeicoes": rejection_reasons,
        "tempo_processamento_s": round(perf_counter() - started, 3),
        "colunas_recebidas": received_columns,
        "colunas_recebidas_normalizadas": normalized_columns,
        "colunas_esperadas": list(staging_columns),
        "colunas_obrigatorias": list(REQUIRED_UPLOAD_COLUMNS),
        "colunas_opcionais": list(staging_columns),
        "colunas_faltantes": [c for c in staging_columns if c not in normalized_columns],
        "colunas_extras": [c for c in normalized_columns if c not in staging_columns],
        "dt_upload": upload_ts.isoformat() if upload_ts else None,
        "observacao_staging": "A staging é temporária e será limpa pela procedure ao final do processamento.",
    }


def _prepare_dataframe_for_staging(
    df,
    filename: str,
    started: float,
    *,
    staging_schema: Optional[List[bigquery.SchemaField]] = None,
    upload_ts: Optional[datetime] = None,
) -> Tuple[Any, Dict[str, Any]]:
    """Normaliza o DataFrame, injeta dt_upload UTC, rejeita linhas vazias e monta relatório."""
    upload_ts = upload_ts or _current_upload_timestamp()
    received_columns = [str(c) for c in df.columns]
    rows_received = len(df)
    df2 = _coerce_df_to_staging_schema(df, staging_schema=staging_schema, upload_ts=upload_ts)
    empty_mask = _empty_row_mask(df2)
    rejected_df = df.loc[empty_mask].copy()
    accepted_df = df2.loc[~empty_mask].copy()

    rejection_reasons: Dict[str, int] = {}
    rejections: List[Dict[str, Any]] = []
    if len(rejected_df) > 0:
        reason = "Linha vazia após normalização"
        rejection_reasons[reason] = len(rejected_df)
        now = datetime.utcnow().isoformat(timespec="seconds")
        for idx, row in rejected_df.iterrows():
            rejections.append(
                {
                    "data_hora": now,
                    "arquivo": filename,
                    "linha": int(idx) + 2,
                    "motivo": reason,
                    "cpf_raw": row.get("cpf"),
                    "celular_raw": row.get("celular"),
                    "nome_raw": row.get("nome"),
                    "email_raw": row.get("email"),
                    "id_importacao": None,
                }
            )
        _insert_rejection_logs(rejections)

    report = _build_import_report(
        filename=filename,
        started=started,
        rows_received=rows_received,
        rows_imported=len(accepted_df),
        rows_rejected=len(rejected_df),
        received_columns=received_columns,
        rejection_reasons=rejection_reasons,
        staging_schema=staging_schema,
        upload_ts=upload_ts,
    )
    return accepted_df, report


def _coerce_df_to_staging_schema(
    df,
    *,
    staging_schema: Optional[List[bigquery.SchemaField]] = None,
    upload_ts: Optional[datetime] = None,
):
    """
    Ajusta o DataFrame para bater com o schema da staging.
    Datas brasileiras permanecem como STRING para a procedure fazer o parse.
    """
    upload_ts = upload_ts or _current_upload_timestamp()

    df = df.copy()
    df.columns = [_normalize_column_name(c) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]

    schema = staging_schema or STAGING_SCHEMA
    expected_cols = [f.name for f in schema]

    for col in expected_cols:
        if col not in df.columns:
            df[col] = None

    # Sempre definido pelo backend; qualquer valor enviado na planilha é ignorado.
    if "dt_upload" in expected_cols:
        df["dt_upload"] = upload_ts
        logger.info("dt_upload aplicado pelo backend: linhas=%d dt_upload=%s", len(df), upload_ts)

    df = df[expected_cols]

    for col in expected_cols:
        if col == "dt_upload":
            continue
        if col in PHONEISH_COLUMNS:
            df[col] = df[col].map(_normalize_phoneish_value)
        else:
            df[col] = df[col].map(_normalize_generic_string)

    for col in expected_cols:
        df[col] = df[col].astype("object")

    return df


# ============================================================
# STAGING + PROCEDURE (upload)  AGORA ASSÍNCRONO
# ============================================================
def _client_identity(client: bigquery.Client) -> str:
    credentials = getattr(client, "_credentials", None)
    if not credentials:
        return "desconhecida"
    return (
        getattr(credentials, "service_account_email", None)
        or getattr(credentials, "signer_email", None)
        or getattr(credentials, "quota_project_id", None)
        or credentials.__class__.__name__
    )


def _assert_staging_has_dt_upload(table: bigquery.Table) -> None:
    if not any(field.name == "dt_upload" and field.field_type.upper() == "TIMESTAMP" for field in table.schema):
        raise RuntimeError(
            "A tabela stg_leads_site não possui a coluna dt_upload TIMESTAMP. "
            "Execute no BigQuery: "
            "ALTER TABLE `painel-universidade.modelo_estrela.stg_leads_site` "
            "ADD COLUMN IF NOT EXISTS dt_upload TIMESTAMP;"
        )


def _ensure_staging_table_exists(client: bigquery.Client, table_id: str) -> bigquery.Table:
    try:
        table = client.get_table(table_id)
    except NotFound as exc:
        raise RuntimeError(
            f"Tabela destino não encontrada no BigQuery: {table_id}. "
            "O upload não cria tabelas automaticamente; crie a tabela de staging antes de enviar arquivos."
        ) from exc

    schema_summary = [f"{field.name}:{field.field_type}:{field.mode}" for field in table.schema]
    _assert_staging_has_dt_upload(table)
    logger.info(
        "Tabela de staging encontrada antes do load: projeto=%s dataset=%s tabela=%s "
        "service_account=%s schema_encontrado=%s total_colunas=%d",
        GCP_PROJECT_ID,
        BQ_DATASET,
        table_id,
        _client_identity(client),
        schema_summary,
        len(table.schema),
    )
    return table


def load_to_staging(
    df,
    *,
    already_coerced: bool = False,
    staging_table: Optional[bigquery.Table] = None,
) -> int:
    client = get_bq_client()
    table_id = _staging_table_id()
    table = staging_table or _ensure_staging_table_exists(client, table_id)
    df2 = df.copy() if already_coerced else _coerce_df_to_staging_schema(df, staging_schema=table.schema)
    write_disposition = bigquery.WriteDisposition.WRITE_TRUNCATE
    create_disposition = bigquery.CreateDisposition.CREATE_NEVER

    def _do_load():
        credentials, project = google.auth.default()
        credential_service_account = getattr(credentials, "service_account_email", "N/A")
        location = _bq_location()
        logger.info(
            "BQ DEBUG | project=%s | cred_type=%s | service_account=%s | table=%s | "
            "dataset=%s | location=%s | write_disposition=%s | create_disposition=%s",
            project,
            credentials.__class__.__name__,
            credential_service_account,
            table_id,
            BQ_DATASET,
            location,
            write_disposition,
            create_disposition,
        )

        table_check = client.get_table(table_id)
        logger.info(
            "BQ TABLE CHECK | project=%s | dataset=%s | table_id=%s | columns=%d | table_location=%s",
            table_check.project,
            table_check.dataset_id,
            table_check.table_id,
            len(table_check.schema),
            getattr(table_check, "location", "N/A"),
        )

        logger.info(
            "Upload iniciado BigQuery: tabela=%s linhas=%d write_disposition=%s create_disposition=%s",
            table_id,
            len(df2),
            write_disposition,
            create_disposition,
        )
        try:
            job = client.load_table_from_dataframe(
                df2,
                table_id,
                job_config=bigquery.LoadJobConfig(
                    schema=table.schema,
                    autodetect=False,
                    create_disposition=create_disposition,
                    write_disposition=write_disposition,
                ),
                location=location,
            )
            job.result()
        except Exception as exc:
            logger.exception(
                "BQ LOAD FAILED | exception_type=%s | message=%s | service_account=%s | table=%s",
                exc.__class__.__name__,
                str(exc),
                credential_service_account,
                table_id,
            )
            raise

        rows_loaded = int(getattr(job, "output_rows", 0) or 0)
        logger.info(
            "Upload concluído BigQuery: job_id=%s tabela=%s linhas_gravadas_staging=%d",
            job.job_id,
            table_id,
            rows_loaded,
        )
        logger.info(
            "A staging é temporária e será limpa pela procedure ao final do processamento: tabela=%s",
            table_id,
        )
        return rows_loaded

    return _with_retry(_do_load, operation_name=f"load_to_staging → {table_id}")


def run_procedure_async() -> str:
    """Dispara a procedure e retorna o job_id sem bloquear a request."""
    client = get_bq_client()
    sql = f"CALL `{GCP_PROJECT_ID}.{BQ_DATASET}.{BQ_PROCEDURE}`();"

    def _do_call():
        job = client.query(sql, location=_bq_location())
        logger.info(
            "Procedure disparada: procedure=%s job_id=%s location=%s",
            BQ_PROCEDURE,
            job.job_id,
            job.location,
        )
        logger.info(
            "Procedure iniciada em modo assíncrono: procedure=%s job_id=%s. Consulte /api/upload/status.",
            BQ_PROCEDURE,
            job.job_id,
        )
        return job.job_id

    return _with_retry(_do_call, operation_name=f"run_procedure_async ({BQ_PROCEDURE})")


def process_upload_dataframe(df, filename: str = "upload") -> Dict[str, Any]:
    """
    1) diagnostica/normaliza o arquivo
    2) carrega a staging oficial (truncate)
    3) dispara SP async
    4) retorna job_id + relatório de importação
    """
    started = perf_counter()
    client = get_bq_client()
    staging_table = _ensure_staging_table_exists(client, _staging_table_id())
    df2, report = _prepare_dataframe_for_staging(df, filename, started, staging_schema=staging_table.schema)
    logger.info(
        "DataFrame preparado para staging: arquivo=%s linhas=%d colunas=%d dt_upload=%s",
        filename,
        len(df2),
        len(df2.columns),
        df2["dt_upload"].iloc[0] if "dt_upload" in df2.columns and len(df2) else None,
    )
    rows_loaded = load_to_staging(df2, already_coerced=True, staging_table=staging_table)
    job_id = run_procedure_async()
    report["linhas_gravadas_staging"] = rows_loaded
    report["procedure_job_id"] = job_id
    report["tempo_processamento_s"] = round(perf_counter() - started, 3)
    logger.info(
        "Processamento BigQuery iniciado: arquivo=%s staging=%s linhas_recebidas=%d "
        "linhas_processadas=%d linhas_gravadas_staging=%d procedure=%s procedure_job_id=%s",
        filename,
        _staging_table_id(),
        report["linhas_recebidas"],
        report["linhas_processadas"],
        rows_loaded,
        BQ_PROCEDURE,
        job_id,
    )
    return {
        "message": "Upload recebido. Dados carregados na staging e procedure iniciada. A staging será limpa ao final da procedure.",
        "job_id": job_id,
        "rows_read": report["linhas_recebidas"],
        "report": report,
    }


def get_bq_job_status(job_id: str) -> Dict[str, Any]:
    client = get_bq_client()
    job = client.get_job(job_id, location=_bq_location())

    payload: Dict[str, Any] = {
        "job_id": job.job_id,
        "state": job.state,
        "location": job.location,
        "created": job.created.isoformat() if job.created else None,
        "started": job.started.isoformat() if job.started else None,
        "ended": job.ended.isoformat() if job.ended else None,
    }

    if job.error_result:
        payload["ok"] = False
        payload["error"] = job.error_result
        payload["errors"] = job.errors
    else:
        payload["ok"] = True if job.state == "DONE" else None
        payload["error"] = None
        payload["errors"] = None

    return payload


# ============================================================
# QUERY HELPERS (sempre a VIEW)
# ============================================================
def _base_select_sql() -> str:
    return f"FROM {_tbl(BQ_VIEW_LEADS)} v WHERE 1=1"


def _apply_filters(sql: str, filters: Dict[str, Any], params: List[Any]) -> str:
    cursos = _as_list(filters.get("curso"))
    polos = _as_list(filters.get("polo"))
    modalidades = _as_list(filters.get("modalidade"))
    turnos = _as_list(filters.get("turno"))
    canais = _as_list(filters.get("canal"))
    campanhas = _as_list(filters.get("campanha"))
    origens = _as_list(filters.get("origem"))
    tipos_negocio = _as_list(filters.get("tipo_negocio"))
    tipos_disparo = _as_list(filters.get("tipo_disparo"))

    status_list = _as_list(filters.get("status"))
    consultores_disp = _as_list(filters.get("consultor_disparo")) or _as_list(filters.get("consultor"))
    consultores_com = _as_list(filters.get("consultor_comercial"))

    sql = _apply_multi_value_filter(sql, params, col="curso", param_name="cursos", values=cursos)
    sql = _apply_multi_value_filter(sql, params, col="polo", param_name="polos", values=polos)
    sql = _apply_multi_value_filter(sql, params, col="modalidade", param_name="modalidades", values=modalidades)
    sql = _apply_multi_value_filter(sql, params, col="turno", param_name="turnos", values=turnos)
    sql = _apply_multi_value_filter(sql, params, col="canal", param_name="canais", values=canais)
    sql = _apply_multi_value_filter(sql, params, col="campanha", param_name="campanhas", values=campanhas)
    sql = _apply_multi_value_filter(sql, params, col="origem", param_name="origens", values=origens)
    sql = _apply_multi_value_filter(sql, params, col="tipo_negocio", param_name="tipos_negocio", values=tipos_negocio)
    sql = _apply_multi_value_filter(sql, params, col="status", param_name="status_list", values=status_list)
    sql = _apply_multi_value_filter(sql, params, col="consultor_disparo", param_name="consultores_disp", values=consultores_disp)
    sql = _apply_multi_value_filter(sql, params, col="consultor_comercial", param_name="consultores_com", values=consultores_com)
    sql = _apply_multi_value_filter(sql, params, col="tipo_disparo", param_name="tipos_disparo", values=tipos_disparo)

    if filters.get("cpf") and _has_view_col("cpf"):
        sql += " AND v.cpf = @cpf"
        params.append(bigquery.ScalarQueryParameter("cpf", "STRING", str(filters["cpf"]).strip()))

    if filters.get("celular") and _has_view_col("celular"):
        sql += " AND v.celular = @celular"
        params.append(bigquery.ScalarQueryParameter("celular", "STRING", str(filters["celular"]).strip()))

    if filters.get("email") and _has_view_col("email"):
        sql += " AND LOWER(v.email) = LOWER(@email)"
        params.append(bigquery.ScalarQueryParameter("email", "STRING", str(filters["email"]).strip()))

    if filters.get("nome") and _has_view_col("nome"):
        sql += " AND LOWER(v.nome) LIKE LOWER(@nome_like)"
        params.append(bigquery.ScalarQueryParameter("nome_like", "STRING", f"%{str(filters['nome']).strip()}%"))

    if filters.get("matriculado") is not None and str(filters.get("matriculado")).strip() != "":
        val = str(filters.get("matriculado")).lower().strip()
        b = True if val in ("true", "1", "sim", "yes") else False if val in ("false", "0", "nao", "não", "no") else None
        if b is not None:
            if _has_view_col("flag_matriculado"):
                sql += " AND IFNULL(v.flag_matriculado, FALSE) = @matriculado"
                params.append(bigquery.ScalarQueryParameter("matriculado", "BOOL", b))
            elif _has_view_col("matriculado"):
                sql += " AND LOWER(CAST(v.matriculado AS STRING)) IN UNNEST(@matriculado_text)"
                truthy = ["true", "1", "sim", "yes", "s"]
                falsy = ["false", "0", "nao", "não", "no", "n"]
                params.append(bigquery.ArrayQueryParameter("matriculado_text", "STRING", truthy if b else falsy))

    if filters.get("data_ini") and _has_view_col("data_inscricao"):
        sql += " AND v.data_inscricao >= @data_ini"
        params.append(bigquery.ScalarQueryParameter("data_ini", "DATE", filters["data_ini"]))

    if filters.get("data_fim") and _has_view_col("data_inscricao"):
        sql += " AND v.data_inscricao <= @data_fim"
        params.append(bigquery.ScalarQueryParameter("data_fim", "DATE", filters["data_fim"]))

    return sql



# ============================================================
# GESTÃO OPERACIONAL (/gestao)
# ============================================================
_SENSITIVE_PARAM_NAME_PARTS = ("cpf", "celular", "telefone", "email", "nome", "busca", "token", "senha", "password", "credential", "payload")


def _format_bq_params_for_log(params: Optional[List[Any]]) -> List[Dict[str, Any]]:
    """Return parameter metadata safe for logs, never raw frontend values."""
    formatted = []
    for param in params or []:
        name = str(getattr(param, "name", "") or "")
        value = getattr(param, "value", None)
        is_sensitive = any(part in name.lower() for part in _SENSITIVE_PARAM_NAME_PARTS)
        formatted.append({
            "name": name or None,
            "type": getattr(param, "type_", None),
            "value": "[REDACTED]" if is_sensitive else ("[SET]" if value not in (None, "") else value),
        })
    return formatted


def _run_gestao_query(sql: str, params: Optional[List[Any]] = None, operation_name: str = "gestão operacional") -> Any:
    client = get_bq_client()
    job_config = bigquery.QueryJobConfig(query_parameters=params or [])
    job = None
    started = perf_counter()
    try:
        logger.info(
            "BQ operation=%s location=%s params=%s",
            operation_name,
            _bq_location(),
            _format_bq_params_for_log(params),
        )
        job = client.query(sql, job_config=job_config, location=_bq_location())
        result = job.result(timeout=QUERY_TIMEOUT_SECONDS)
    except FuturesTimeoutError as exc:
        logger.exception("BQ operation=%s timeout job_id=%s", operation_name, getattr(job, "job_id", None))
        if job is not None:
            job.cancel()
        raise TimeoutError(f"Consulta excedeu timeout de {QUERY_TIMEOUT_SECONDS}s") from exc
    except Exception:
        logger.exception(
            "BQ operation=%s falhou job_id=%s errors=%s params=%s",
            operation_name,
            getattr(job, "job_id", None),
            getattr(job, "errors", None),
            _format_bq_params_for_log(params),
        )
        raise
    logger.info("BQ %s concluído job_id=%s duration_ms=%s", operation_name, getattr(job, "job_id", None), int((perf_counter() - started) * 1000))
    return result


def query_gestao_operacao() -> Dict[str, Any]:
    """Consulta KPIs consolidados com prioridade baseada exclusivamente em status."""
    status_empty_expr = _gestao_status_empty_expr("v")
    matriculado_expr = _gestao_matriculado_expr("v")
    data_inscricao_expr = _gestao_timestamp_expr("data_inscricao", alias="v")
    primeiro_contato_expr = _gestao_timestamp_expr("data_primeiro_contato", "primeiro_contato", "data_disparo", alias="v")
    matricula_data_expr = _gestao_timestamp_expr("data_matricula", "data_matriculado", "dt_matricula", alias="v")
    score_expr = "SAFE_CAST(v.score_prioridade AS FLOAT64)" if _has_view_col("score_prioridade") else f"CASE WHEN {matriculado_expr} THEN 10 WHEN {status_empty_expr} THEN 100 ELSE 50 END"
    sql = f"""
    SELECT
      COUNT(*) AS total_leads,
      COUNTIF({status_empty_expr}) AS nunca_trabalhados,
      COUNTIF({status_empty_expr}) AS aguardando_primeiro_contato,
      COUNTIF(NOT {status_empty_expr} AND NOT {matriculado_expr}) AS leads_em_carteira,
      COUNTIF({matriculado_expr}) AS matriculados,
      COUNTIF({status_empty_expr}) AS leads_criticos,
      COUNTIF({status_empty_expr}) AS leads_alta_prioridade,
      COUNTIF({status_empty_expr}) AS sem_acionamento,
      COUNTIF({primeiro_contato_expr} IS NULL) AS leads_parados,
      SAFE_DIVIDE(COUNTIF({matriculado_expr}), COUNT(*)) * 100 AS taxa_geral_conversao,
      AVG(IF({primeiro_contato_expr} IS NOT NULL AND {data_inscricao_expr} IS NOT NULL, TIMESTAMP_DIFF({primeiro_contato_expr}, {data_inscricao_expr}, HOUR), NULL)) AS media_horas_primeiro_contato,
      AVG(IF({matricula_data_expr} IS NOT NULL AND {data_inscricao_expr} IS NOT NULL, TIMESTAMP_DIFF({matricula_data_expr}, {data_inscricao_expr}, HOUR), NULL)) AS media_horas_ate_matricula,
      AVG({score_expr}) AS score_medio_operacao,
      COUNTIF({status_empty_expr}) AS validacao_nunca_trabalhados_count,
      COUNTIF({status_empty_expr}) AS validacao_nunca_trabalhados_card,
      COUNTIF({status_empty_expr}) - 123301 AS validacao_nunca_trabalhados_diferenca
    FROM {_tbl(BQ_VIEW_LEADS)} v
    """
    return _single_row_to_json_safe(_run_gestao_query(sql, operation_name="query_gestao_operacao"))

def _gestao_timestamp_expr(*candidate_cols: str, alias: str = "v") -> str:
    col = _first_existing_col(*candidate_cols)
    if not col:
        return "CAST(NULL AS TIMESTAMP)"
    raw = f"TRIM(CAST({alias}.{col} AS STRING))"
    return (
        f"COALESCE("
        f"SAFE_CAST({alias}.{col} AS TIMESTAMP), "
        f"TIMESTAMP(SAFE_CAST({alias}.{col} AS DATE)), "
        f"TIMESTAMP(SAFE.PARSE_DATE('%Y-%m-%d', SUBSTR({raw}, 1, 10))), "
        f"TIMESTAMP(SAFE.PARSE_DATE('%d/%m/%Y', SUBSTR({raw}, 1, 10))), "
        f"SAFE.PARSE_TIMESTAMP('%Y-%m-%d %H:%M:%S', SUBSTR({raw}, 1, 19)), "
        f"SAFE.PARSE_TIMESTAMP('%d/%m/%Y %H:%M:%S', SUBSTR({raw}, 1, 19)), "
        f"IF(SAFE_CAST({raw} AS INT64) IS NULL, NULL, TIMESTAMP(DATE_ADD(DATE '1899-12-30', INTERVAL SAFE_CAST({raw} AS INT64) DAY)))"
        f")"
    )


def _gestao_date_expr(*candidate_cols: str, alias: str = "v") -> str:
    ts = _gestao_timestamp_expr(*candidate_cols, alias=alias)
    return f"DATE({ts})"


def _gestao_text_expr(*candidate_cols: str, alias: str = "v", default: str = "Não informado") -> str:
    col = _first_existing_col(*candidate_cols)
    if not col:
        return f"'{default}'"
    return f"COALESCE(NULLIF(TRIM(CAST({alias}.{col} AS STRING)), ''), '{default}')"


def query_gestao_produtividade() -> List[Dict[str, Any]]:
    """Consulta produtividade por consultor com conversão MAT/total diretamente na base da gestão."""
    status_empty_expr = _gestao_status_empty_expr("v")
    matriculado_expr = _gestao_matriculado_expr("v")
    consultor_expr = _gestao_text_expr("consultor_comercial", alias="v", default="Sem consultor")
    data_inscricao_expr = _gestao_timestamp_expr("data_inscricao", alias="v")
    primeiro_contato_expr = _gestao_timestamp_expr("data_primeiro_contato", "primeiro_contato", "data_disparo", alias="v")
    matricula_data_expr = _gestao_timestamp_expr("data_matricula", "data_matriculado", "dt_matricula", alias="v")
    ultima_atividade_expr = _gestao_timestamp_expr("data_ultima_acao", "ultima_atividade", "data_disparo", alias="v")
    score_expr = "SAFE_CAST(v.score_prioridade AS FLOAT64)" if _has_view_col("score_prioridade") else f"CASE WHEN {matriculado_expr} THEN 10 WHEN {status_empty_expr} THEN 100 ELSE 50 END"
    sql = f"""
    SELECT
      {consultor_expr} AS consultor_comercial,
      COUNT(*) AS total_leads,
      COUNTIF({status_empty_expr}) AS leads_sem_status,
      COUNTIF(NOT {status_empty_expr} AND NOT {matriculado_expr}) AS leads_em_carteira,
      COUNTIF({matriculado_expr}) AS matriculados,
      SAFE_DIVIDE(COUNTIF({matriculado_expr}), COUNT(*)) * 100 AS taxa_conversao_pct,
      SAFE_DIVIDE(COUNTIF({matriculado_expr}), COUNT(*)) * 100 AS taxa_matricula_pct,
      AVG(IF({primeiro_contato_expr} IS NOT NULL AND {data_inscricao_expr} IS NOT NULL, TIMESTAMP_DIFF({primeiro_contato_expr}, {data_inscricao_expr}, HOUR), NULL)) AS media_horas_primeiro_contato,
      AVG(IF({matricula_data_expr} IS NOT NULL AND {data_inscricao_expr} IS NOT NULL, TIMESTAMP_DIFF({matricula_data_expr}, {data_inscricao_expr}, HOUR), NULL)) AS media_horas_ate_matricula,
      MAX({ultima_atividade_expr}) AS ultima_atividade,
      AVG({score_expr}) AS score_medio_carteira,
      COUNTIF({ultima_atividade_expr} IS NULL OR DATE_DIFF(CURRENT_DATE(), DATE({ultima_atividade_expr}), DAY) > 7) AS leads_sem_movimento_7_dias,
      COUNTIF({status_empty_expr}) AS leads_criticos,
      COUNTIF({status_empty_expr}) AS leads_alta_prioridade,
      COUNTIF({status_empty_expr}) AS leads_nao_disparados,
      COUNTIF(NOT {status_empty_expr}) AS leads_disparados
    FROM {_tbl(BQ_VIEW_LEADS)} v
    GROUP BY consultor_comercial
    ORDER BY taxa_conversao_pct DESC, matriculados DESC, total_leads DESC
    """
    return _rows_to_json_safe(_run_gestao_query(sql, operation_name="query_gestao_produtividade"))

def query_gestao_fila_operacional(limit: int = GESTAO_FILA_LIMIT) -> List[Dict[str, Any]]:
    """Consulta fila priorizada recalculando prioridade sem data_disparo."""
    limit = max(1, min(int(limit or GESTAO_FILA_LIMIT), 5000))
    columns = _view_columns()
    status_expr = "v.status" if "status" in columns else "v.status_inscricao" if "status_inscricao" in columns else "CAST(NULL AS STRING)"
    polo_expr = "v.polo" if "polo" in columns else "v.unidade" if "unidade" in columns else "CAST(NULL AS STRING)"
    status_empty_expr = _gestao_status_empty_expr("v")
    matriculado_expr = _gestao_matriculado_expr("v")
    ultima_acao_expr = "SAFE_CAST(v.data_ultima_acao AS TIMESTAMP)" if "data_ultima_acao" in columns else "CAST(NULL AS TIMESTAMP)"
    data_inscricao_expr = "SAFE_CAST(v.data_inscricao AS DATE)" if "data_inscricao" in columns else "CAST(NULL AS DATE)"
    nunca_disparado_expr = "(v.data_disparo IS NULL OR TRIM(CAST(v.data_disparo AS STRING)) = '')" if "data_disparo" in columns else "CAST(NULL AS BOOL)"
    dias_sem_acao_expr = f"IF({ultima_acao_expr} IS NULL, NULL, DATE_DIFF(CURRENT_DATE(), DATE({ultima_acao_expr}), DAY))"
    nivel_prioridade_expr = f"""
        CASE
          WHEN {matriculado_expr} THEN 'BAIXA'
          WHEN {status_empty_expr} THEN 'CRÍTICA'
          ELSE 'MÉDIA'
        END
    """
    score_expr = f"""
        CASE
          WHEN {matriculado_expr} THEN 10
          WHEN {status_empty_expr} THEN 100
          ELSE 50
        END + COALESCE({dias_sem_acao_expr}, 0) * 0.01
    """
    etapa_expr = f"""
        CASE
          WHEN {matriculado_expr} THEN 'MATRICULADO'
          WHEN {status_empty_expr} THEN 'NUNCA_TRABALHADO'
          ELSE 'EM_CARTEIRA'
        END
    """
    select_cols = ",\n      ".join([
        _select_col("nome"),
        _select_col("celular"),
        _select_col("curso"),
        f"{polo_expr} AS polo",
        _select_col("origem"),
        f"{status_expr} AS status",
        _select_col("consultor_comercial"),
        f"{score_expr} AS score_prioridade",
        f"{nivel_prioridade_expr} AS nivel_prioridade",
        f"{etapa_expr} AS etapa_operacional",
        f"{dias_sem_acao_expr} AS dias_sem_acao",
        f"{nunca_disparado_expr} AS nunca_disparado",
        "CAST(NULL AS FLOAT64) AS horas_ate_primeiro_contato",
        f"{matriculado_expr} AS flag_matriculado",
        f"{data_inscricao_expr} AS data_inscricao",
        f"{ultima_acao_expr} AS data_ultima_acao",
    ])
    logger.info("Colunas disponíveis em %s: %s", _view_table_id(), sorted(columns))
    sql = f"""
    SELECT *
    FROM (
      SELECT
        {select_cols}
      FROM {_tbl(BQ_VIEW_LEADS)} v
    )
    ORDER BY COALESCE(score_prioridade, 0) DESC, data_inscricao DESC
    LIMIT @limit
    """
    params = [bigquery.ScalarQueryParameter("limit", "INT64", limit)]
    return _rows_to_json_safe(_run_gestao_query(sql, params=params, operation_name="query_gestao_fila_operacional"))

def _gestao_number(value: Any) -> float:
    if value is None or value == "":
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0



def _gestao_status_empty_expr(alias: str = "v") -> str:
    """Status vazio normalizado: NULL, string vazia e marcadores textuais de ausência."""
    if _has_view_col("status"):
        return (
            f"(NULLIF(TRIM(CAST({alias}.status AS STRING)), '') IS NULL OR "
            f"UPPER(TRIM(CAST({alias}.status AS STRING))) IN ('NULL','N/A','NA','SEM STATUS','SEM INFORMACAO','SEM INFORMAÇÃO','-'))"
        )
    return "FALSE"


def _gestao_matriculado_expr(alias: str = "v") -> str:
    """Regra oficial de matrícula: status textual ou flag_matriculado = TRUE."""
    checks = []
    if _has_view_col("status"):
        checks.append(f"UPPER(TRIM(CAST({alias}.status AS STRING))) IN ('MAT','MATRICULADO')")
    if _has_view_col("flag_matriculado"):
        checks.append(f"COALESCE(SAFE_CAST({alias}.flag_matriculado AS BOOL), FALSE)")
    if _has_view_col("matriculado"):
        checks.append(
            f"UPPER(TRIM(CAST({alias}.matriculado AS STRING))) IN "
            "('SIM', 'S', 'TRUE', '1', 'MATRICULADO', 'MAT')"
        )
    return "(" + " OR ".join(checks) + ")" if checks else "FALSE"


def _gestao_criticos_por_consultor_cte() -> str:
    if not _has_view_col("consultor_comercial"):
        return "SELECT CAST(NULL AS STRING) AS consultor_key, 0 AS leads_criticos WHERE FALSE"
    return f"""
      SELECT
        COALESCE(NULLIF(TRIM(CAST(v.consultor_comercial AS STRING)), ''), 'Sem consultor') AS consultor_key,
        COUNTIF({_gestao_status_empty_expr('v')}) AS leads_criticos
      FROM {_tbl(BQ_VIEW_LEADS)} v
      GROUP BY consultor_key
    """

def _gestao_recent_activity_alert_count(produtividade: List[Dict[str, Any]]) -> int:
    threshold = datetime.utcnow() - timedelta(days=7)
    count = 0
    for row in produtividade:
        raw = row.get("ultima_atividade")
        if not raw:
            count += 1
            continue
        parsed = None
        if isinstance(raw, datetime):
            parsed = raw.replace(tzinfo=None)
        else:
            text = str(raw).replace("Z", "+00:00")
            try:
                parsed = datetime.fromisoformat(text).replace(tzinfo=None)
            except ValueError:
                try:
                    parsed = datetime.strptime(str(raw)[:10], "%Y-%m-%d")
                except ValueError:
                    parsed = None
        if parsed is None or parsed < threshold:
            count += 1
    return count


def query_gestao_exportar_prioritarios(limit: int = 500) -> Tuple[str, bytes, int]:
    """Gera CSV em memória com leads prioritários status IS NULL."""
    allowed_limits = {100, 500, 1000, 5000}
    limit = int(limit or 500)
    if limit not in allowed_limits:
        limit = 500
    columns = _view_columns()
    status_empty_expr = _gestao_status_empty_expr("v")
    polo_expr = "v.polo" if "polo" in columns else "v.unidade" if "unidade" in columns else "CAST(NULL AS STRING)"
    data_inscricao_expr = _gestao_date_expr("data_inscricao", alias="v")
    score_expr = "SAFE_CAST(v.score_prioridade AS FLOAT64)" if "score_prioridade" in columns else "100"
    nivel_expr = "v.nivel_prioridade" if "nivel_prioridade" in columns else "'CRÍTICA'"
    sql = f"""
    SELECT
      {_select_col('nome')},
      {_select_col('celular')},
      {_select_col('email')},
      {_select_col('curso')},
      {polo_expr} AS polo,
      {_select_col('origem')},
      {_select_col('consultor_comercial')},
      {data_inscricao_expr} AS data_inscricao,
      {score_expr} AS score_prioridade,
      {nivel_expr} AS nivel_prioridade
    FROM {_tbl(BQ_VIEW_LEADS)} v
    WHERE {status_empty_expr}
    ORDER BY score_prioridade DESC, data_inscricao DESC
    LIMIT @limit
    """
    params = [bigquery.ScalarQueryParameter("limit", "INT64", limit)]
    rows = _rows_to_json_safe(_run_gestao_query(sql, params=params, operation_name="query_gestao_exportar_prioritarios"))
    output = StringIO()
    fieldnames = ["nome", "celular", "email", "curso", "polo", "origem", "consultor_comercial", "data_inscricao", "score_prioridade", "nivel_prioridade"]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    filename = f"leads_prioritarios_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
    logger.info("Gestão exportação prioritários gerada filename=%s limit=%s rows=%s", filename, limit, len(rows))
    return filename, output.getvalue().encode("utf-8-sig"), len(rows)


def query_gestao_dimensao(campo: str, limit: int = 10) -> List[Dict[str, Any]]:
    """Ranking por dimensão de volume/conversão usado em origem e curso."""
    allowed = {"origem": ("origem",), "curso": ("curso",)}
    if campo not in allowed:
        return []
    dim_expr = _gestao_text_expr(*allowed[campo], alias="v", default="Não informado")
    matriculado_expr = _gestao_matriculado_expr("v")
    sql = f"""
    SELECT
      {dim_expr} AS dimensao,
      COUNT(*) AS total_leads,
      COUNTIF({matriculado_expr}) AS matriculados,
      SAFE_DIVIDE(COUNTIF({matriculado_expr}), COUNT(*)) * 100 AS taxa_conversao_pct
    FROM {_tbl(BQ_VIEW_LEADS)} v
    GROUP BY dimensao
    HAVING total_leads > 0
    ORDER BY total_leads DESC
    LIMIT @limit
    """
    params = [bigquery.ScalarQueryParameter("limit", "INT64", limit)]
    return _rows_to_json_safe(_run_gestao_query(sql, params=params, operation_name=f"query_gestao_dimensao_{campo}"))


def query_gestao_qualidade_dados() -> Dict[str, Any]:
    """Indicadores de qualidade de dados para a gestão."""
    telefone_expr = _gestao_text_expr("celular", "telefone", alias="v", default="")
    email_expr = _gestao_text_expr("email", alias="v", default="")
    origem_expr = _gestao_text_expr("origem", alias="v", default="")
    curso_expr = _gestao_text_expr("curso", alias="v", default="")
    cpf_col = _first_existing_col("cpf")
    cpf_dup = f"COUNT(*) - COUNT(DISTINCT NULLIF(TRIM(CAST(v.{cpf_col} AS STRING)), ''))" if cpf_col else "0"
    tel_col = _first_existing_col("celular", "telefone")
    tel_dup = f"COUNT(*) - COUNT(DISTINCT NULLIF(REGEXP_REPLACE(CAST(v.{tel_col} AS STRING), r'\\D', ''), ''))" if tel_col else "0"
    sql = f"""
    SELECT
      COUNTIF({telefone_expr} = '') AS leads_sem_telefone,
      COUNTIF({email_expr} = '') AS leads_sem_email,
      {cpf_dup} AS leads_duplicados_cpf,
      {tel_dup} AS leads_duplicados_telefone,
      COUNTIF({origem_expr} = '') AS leads_sem_origem,
      COUNTIF({curso_expr} = '') AS leads_sem_curso
    FROM {_tbl(BQ_VIEW_LEADS)} v
    """
    return _single_row_to_json_safe(_run_gestao_query(sql, operation_name="query_gestao_qualidade_dados"))


def _gestao_melhor(rows: List[Dict[str, Any]], label_key: str = "dimensao") -> str:
    ranked = sorted(rows, key=lambda r: (_gestao_number(r.get("taxa_conversao_pct")), _gestao_number(r.get("matriculados"))), reverse=True)
    return str(ranked[0].get(label_key) or "-") if ranked else "-"


def _build_gestao_alertas(operacao: Dict[str, Any], produtividade: List[Dict[str, Any]]) -> Dict[str, Any]:
    leads_sem_movimento = sum(_gestao_number(row.get("leads_sem_movimento_7_dias")) for row in produtividade)
    if leads_sem_movimento == 0:
        leads_sem_movimento = _gestao_number(operacao.get("leads_parados"))

    total_carteira = sum(_gestao_number(row.get("leads_em_carteira")) for row in produtividade)
    media_carteira = total_carteira / len(produtividade) if produtividade else 0
    acima_media = sum(1 for row in produtividade if _gestao_number(row.get("leads_em_carteira")) > media_carteira)
    melhor_conversao = max((_gestao_number(row.get("taxa_conversao_pct")) for row in produtividade), default=0)

    return {
        "leads_sem_status": int(_gestao_number(operacao.get("nunca_trabalhados"))),
        "leads_sem_acao_7_dias": int(leads_sem_movimento),
        "consultores_sem_atividade_recente": _gestao_recent_activity_alert_count(produtividade),
        "consultores_carteira_acima_media": int(acima_media),
        "consultores_melhor_conversao": int(sum(1 for row in produtividade if melhor_conversao and _gestao_number(row.get("taxa_conversao_pct")) == melhor_conversao)),
        "leads_criticos": int(_gestao_number(operacao.get("leads_criticos"))),
    }


def query_gestao_dashboard(force_refresh: bool = False) -> Dict[str, Any]:
    """Retorna todos os dados do módulo /gestao com cache curto para evitar consultas repetidas."""
    cache_key = f"gestao_dashboard::{GCP_PROJECT_ID}.{BQ_DATASET}::{GESTAO_FILA_LIMIT}"
    with _gestao_cache_lock:
        if not force_refresh and cache_key in _gestao_cache:
            cached = _gestao_cache[cache_key]
            logger.info("Gestão Operacional servida do cache ttl=%ss", GESTAO_CACHE_TTL_SECONDS)
            return cached

    started = perf_counter()
    logger.info("Gestão Operacional iniciando query_gestao_operacao")
    operacao = query_gestao_operacao()
    logger.info("Gestão Operacional query_gestao_operacao concluída")
    logger.info("Gestão Operacional iniciando query_gestao_produtividade")
    produtividade = query_gestao_produtividade()
    logger.info("Gestão Operacional query_gestao_produtividade concluída rows=%s", len(produtividade))
    logger.info("Gestão Operacional iniciando query_gestao_fila_operacional limit=%s", GESTAO_FILA_LIMIT)
    fila = query_gestao_fila_operacional(limit=GESTAO_FILA_LIMIT)
    logger.info("Gestão Operacional query_gestao_fila_operacional concluída rows=%s", len(fila))
    top_origens = query_gestao_dimensao("origem", limit=10)
    top_cursos = query_gestao_dimensao("curso", limit=10)
    qualidade_dados = query_gestao_qualidade_dados()
    consultor_campeao = (produtividade[0].get("consultor_comercial") if produtividade else "-")
    centro_comando = {
        "total_leads": operacao.get("total_leads"),
        "nunca_trabalhados": operacao.get("nunca_trabalhados"),
        "leads_em_carteira": operacao.get("leads_em_carteira"),
        "matriculados": operacao.get("matriculados"),
        "taxa_geral_conversao": operacao.get("taxa_geral_conversao"),
        "media_horas_primeiro_contato": operacao.get("media_horas_primeiro_contato"),
        "media_horas_ate_matricula": operacao.get("media_horas_ate_matricula"),
        "consultor_campeao_dia": consultor_campeao,
        "origem_melhor_conversao": _gestao_melhor(top_origens),
        "curso_melhor_conversao": _gestao_melhor(top_cursos),
    }
    payload = {
        "operacao": operacao,
        "produtividade": produtividade,
        "fila_operacional": fila,
        "alertas": _build_gestao_alertas(operacao, produtividade),
        "top_origens": top_origens,
        "top_cursos": top_cursos,
        "centro_comando": centro_comando,
        "qualidade_dados": qualidade_dados,
        "cache_ttl_seconds": GESTAO_CACHE_TTL_SECONDS,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }

    with _gestao_cache_lock:
        _gestao_cache[cache_key] = payload

    logger.info(
        "Gestão Operacional carregada consultores=%s fila=%s elapsed=%.2fs",
        len(produtividade),
        len(fila),
        perf_counter() - started,
    )
    return payload


# ============================================================
# LISTAGEM
# ============================================================
def query_leads(
    filters: Optional[Dict[str, Any]] = None,
    limit: int = DEFAULT_LIMIT,
    offset: int = 0,
    order_by: str = "data_inscricao",
    order_dir: str = "DESC",
) -> List[Dict[str, Any]]:
    return list(query_leads_iter(filters=filters, limit=limit, offset=offset, order_by=order_by, order_dir=order_dir))


def query_leads_iter(
    filters: Optional[Dict[str, Any]] = None,
    limit: int = DEFAULT_LIMIT,
    offset: int = 0,
    order_by: str = "data_inscricao",
    order_dir: str = "DESC",
) -> Iterator[Dict[str, Any]]:
    client = get_bq_client()
    filters = filters or {}

    limit = max(1, min(int(limit), MAX_LIMIT))
    offset = max(0, int(offset))
    order_dir = "ASC" if str(order_dir).upper() == "ASC" else "DESC"

    if order_by == "data_inscricao_dt":
        order_by = "data_inscricao"

    order_expr = _order_expr_for(order_by)
    select_cols = ",\n      ".join([
        _select_col("data_inscricao", bq_type="DATE"),
        _select_col("nome"),
        _select_col("cpf"),
        _select_col("celular"),
        _select_col("email"),
        _select_col("curso"),
        _select_col("modalidade"),
        _select_col("turno"),
        _select_col("polo"),
        _select_col("origem"),
        _select_col("status"),
        _select_col("flag_matriculado", bq_type="BOOL"),
        _select_col("consultor_comercial"),
        _select_col("consultor_disparo"),
        _select_col("canal"),
        _select_col("campanha"),
    ])

    sql = f"""
    SELECT
      {select_cols}
    """ + _base_select_sql()

    params: List[Any] = []
    sql = _apply_filters(sql, filters, params)
    if order_by in ("data_inscricao", "data_inscricao_dt"):
        order_clause = _data_inscricao_order_clause(order_dir)
    elif order_by == "data_disparo":
        order_clause = _data_disparo_order_clause(order_dir)
    elif order_expr:
        order_clause = f"{order_expr} {order_dir}"
    else:
        order_clause = "1"
    sql += f"\n ORDER BY {order_clause} \n LIMIT @limit OFFSET @offset"

    params.append(bigquery.ScalarQueryParameter("limit", "INT64", limit))
    params.append(bigquery.ScalarQueryParameter("offset", "INT64", offset))

    t0 = perf_counter()
    job = client.query(sql, job_config=bigquery.QueryJobConfig(query_parameters=params))
    try:
        rows = job.result(timeout=QUERY_TIMEOUT_SECONDS)
    except FuturesTimeoutError as exc:
        logger.exception("BQ query_leads_iter timeout job_id=%s", getattr(job, "job_id", None))
        job.cancel()
        raise TimeoutError(f"Consulta excedeu timeout de {QUERY_TIMEOUT_SECONDS}s") from exc
    logger.info("BQ query_leads_iter concluído job_id=%s elapsed=%.2fs", getattr(job, "job_id", None), perf_counter() - t0)
    for r in rows:
        yield dict(r)


def query_leads_count(filters: Optional[Dict[str, Any]] = None) -> int:
    client = get_bq_client()
    filters = filters or {}

    sql = "SELECT COUNT(1) AS total " + _base_select_sql()
    params: List[Any] = []
    sql = _apply_filters(sql, filters, params)

    rows = list(client.query(sql, job_config=bigquery.QueryJobConfig(query_parameters=params)).result())
    return int(rows[0]["total"]) if rows else 0


# ============================================================
# OPTIONS
# ============================================================
def _distinct_values_from_view(col: str, alias: str) -> List[str]:
    if not _has_view_col(col):
        logger.warning("Coluna '%s' não existe em %s; opções de filtro vazias.", col, _view_table_id())
        return []

    client = get_bq_client()
    sql = f"""
    SELECT DISTINCT {col} AS {alias}
    FROM {_tbl(BQ_VIEW_LEADS)}
    WHERE {col} IS NOT NULL AND TRIM(CAST({col} AS STRING)) != ''
    ORDER BY {alias}
    """
    return [str(r[alias]) for r in client.query(sql).result()]


def query_options() -> Dict[str, List[str]]:
    return {
        "status": _distinct_values_from_view("status", "status"),
        "cursos": _distinct_values_from_view("curso", "curso"),
        "modalidades": _distinct_values_from_view("modalidade", "modalidade"),
        "turnos": _distinct_values_from_view("turno", "turno"),
        "polos": _distinct_values_from_view("polo", "polo"),
        "origens": _distinct_values_from_view("origem", "origem"),
        "canais": _distinct_values_from_view("canal", "canal"),
        "campanhas": _distinct_values_from_view("campanha", "campanha"),
        "consultores_disparo": _distinct_values_from_view("consultor_disparo", "consultor_disparo"),
        "consultores_comercial": _distinct_values_from_view("consultor_comercial", "consultor_comercial"),
        "tipos_disparo": _distinct_values_from_view("tipo_disparo", "tipo_disparo"),
        "tipos_negocio": _distinct_values_from_view("tipo_negocio", "tipo_negocio"),
    }


# ============================================================
# EXPORT (XLSX)
# ============================================================
EXPORT_COLUMNS: List[Tuple[str, str]] = [
    ("status_inscricao", "status_inscricao"),
    ("data_inscricao", "data_inscricao"),
    ("origem", "origem"),
    ("unidade", "unidade"),
    ("tipo_negocio", "tipo_negocio"),
    ("curso", "curso"),
    ("modalidade", "modalidade"),
    ("turno", "turno"),
    ("nome", "nome"),
    ("cpf", "cpf"),
    ("celular", "celular"),
    ("email", "email"),
    ("data_ultima_acao", "data_ultima_acao"),
    ("qtd_acionamentos", "qtd_acionamentos"),
    ("status", "status"),
    ("data_disparo", "data_disparo"),
    ("peca_disparo", "peca_disparo"),
    ("texto_disparo", "texto_disparo"),
    ("consultor_disparo", "consultor_disparo"),
    ("tipo_disparo", "tipo_disparo"),
    ("campanha", "campanha"),
    ("observacao", "observacao"),
    ("data_matricula", "data_matricula"),
    ("matriculado", "matriculado"),
    ("canal", "canal"),
    ("acao_comercial", "acao_comercial"),
    ("consultor_comercial", "consultor_comercial"),
]

EXPORT_COLUMN_FALLBACKS: Dict[str, Tuple[str, ...]] = {
    "unidade": ("unidade", "polo"),
    "matriculado": ("matriculado", "flag_matriculado"),
}


def _select_export_col(col: str, bq_type: str = "STRING") -> str:
    """Seleciona coluna de exportação preservando o alias esperado no cabeçalho."""
    for source_col in EXPORT_COLUMN_FALLBACKS.get(col, (col,)):
        if _has_view_col(source_col):
            return f"v.{source_col} AS {col}"
    return f"CAST(NULL AS {bq_type}) AS {col}"


def export_leads_rows(
    filters: Optional[Dict[str, Any]] = None,
    limit: int = EXPORT_MAX_ROWS,
    offset: int = 0,
    order_by: str = "data_inscricao",
    order_dir: str = "DESC",
) -> List[Dict[str, Any]]:
    client = get_bq_client()
    filters = filters or {}

    limit = max(1, min(int(limit), EXPORT_MAX_ROWS))
    offset = max(0, int(offset))
    order_dir = "ASC" if str(order_dir).upper() == "ASC" else "DESC"

    if order_by == "data_inscricao_dt":
        order_by = "data_inscricao"

    order_expr = _order_expr_for(order_by)
    bool_cols = {"flag_matriculado"}
    date_cols = {"data_inscricao", "data_ultima_acao", "data_disparo", "data_matricula"}
    select_cols = ",\n      ".join([
        _select_export_col(c, bq_type=("BOOL" if c in bool_cols else "DATE" if c in date_cols else "STRING"))
        for c, _ in EXPORT_COLUMNS
    ])

    sql = f"""
    SELECT
      {select_cols}
    """ + _base_select_sql()

    params: List[Any] = []
    sql = _apply_filters(sql, filters, params)

    if order_by in ("data_inscricao", "data_inscricao_dt"):
        order_clause = _data_inscricao_order_clause(order_dir)
    elif order_by == "data_disparo":
        order_clause = _data_disparo_order_clause(order_dir)
    elif order_expr:
        order_clause = f"{order_expr} {order_dir}"
    else:
        order_clause = "1"
    sql += f"\n ORDER BY {order_clause} \n LIMIT @limit OFFSET @offset"
    params.append(bigquery.ScalarQueryParameter("limit", "INT64", limit))
    params.append(bigquery.ScalarQueryParameter("offset", "INT64", offset))

    rows = client.query(sql, job_config=bigquery.QueryJobConfig(query_parameters=params), location=_bq_location()).result()
    return [dict(r) for r in rows]


def export_leads_rows_iter(
    filters: Optional[Dict[str, Any]] = None,
    batch_size: int = 1000,
    order_by: str = "data_inscricao",
    order_dir: str = "DESC",
) -> Iterator[List[Dict[str, Any]]]:
    """Itera exportação paginada para evitar alto consumo de memória."""
    offset = 0
    size = max(1, int(batch_size))
    while True:
        rows = export_leads_rows(filters=filters, limit=size, offset=offset, order_by=order_by, order_dir=order_dir)
        if not rows:
            break
        yield rows
        fetched = len(rows)
        offset += fetched
        if fetched < size:
            break


def rows_to_xlsx(rows: List[Dict[str, Any]], xlsx_path: str, sheet_name: str = "Leads") -> str:
    """
    Gera XLSX no disco.
    - Excel não aceita datetime com timezone
    - CPF/celular saem como TEXTO para evitar conversão numérica
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = [label for _, label in EXPORT_COLUMNS]
    keys = [key for key, _ in EXPORT_COLUMNS]
    ws.append(headers)

    for r in rows:
        ws.append([_xlsx_safe_cell_value(k, r.get(k)) for k in keys])

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(12, min(42, len(str(header)) + 6))

    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


def df_to_xlsx(df, xlsx_path: str, sheet_name: str = "Upload") -> str:
    """
    Salva uma cópia do upload em XLSX.
    Mantém CPF/celular como texto quando possível.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = [str(c) for c in df.columns]
    ws.append(headers)

    phoneish_indexes = {idx for idx, col in enumerate(headers) if str(col).strip() in PHONEISH_COLUMNS}

    for row in df.itertuples(index=False, name=None):
        out = []
        for idx, value in enumerate(row):
            col_name = headers[idx]
            if idx in phoneish_indexes or col_name in PHONEISH_COLUMNS:
                out.append(_normalize_phoneish_value(value))
            else:
                out.append(value)
        ws.append(out)

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(12, min(42, len(str(header)) + 6))

    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path
